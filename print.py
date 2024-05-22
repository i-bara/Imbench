import os
import re
import json
import datetime
import pandas as pd
from openpyxl import load_workbook
import statistics

method_list = ['vanilla', 'drgcn', 'smote', 'imgagn', 'ens', 'tam', 'lte4g', 'topoauc', 'sann', 'sha', 'renode', 'pastel', 'hyperimba']
score_list = ['acc', 'bacc', 'f1', 'auc']
dataset_list = ['Cora_100', 'Cora_20', 'Cora_1', 
                          'CiteSeer_100', 'CiteSeer_20', 'CiteSeer_1', 
                          'PubMed_100', 'PubMed_20', 'PubMed_1', 
                          'chameleon_100', 'chameleon_20', 'chameleon_1', 
                          'squirrel_100', 'squirrel_20', 'squirrel_1', 
                          'Actor_100', 'Actor_20', 'Actor_1', 
                          'Wisconsin_100', 'Wisconsin_20', 'Wisconsin_1', 
                          'Amazon-Photo_100', 'Amazon-Photo_20', 'Amazon-Photo_1', 
                          'Amazon-Computers_100', 'Amazon-Computers_20', 'Amazon-Computers_1', 
                          'ogbn-arxiv_100', 'ogbn-arxiv_20', 'ogbn-arxiv_1', 
                          'Amazon-Photo', 'Amazon-Computers', 'Coauthor-CS', 'ogbn-arxiv']
seed_list = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]

def get_method_offset(method):
    if method == 'vanilla':
        return 0
    elif method == 'drgcn':
        return 1
    elif method == 'smote':
        return 2
    elif method == 'imgagn':
        return 3
    elif method == 'ens':
        return 4
    elif method == 'tam':
        return 5
    elif method == 'lte4g':
        return 6
    elif method == 'topoauc':
        return 7
    elif method == 'sann':
        return 8
    elif method == 'sha':
        return 9
    elif method == 'renode':
        return 11
    elif method == 'pastel':
        return 12
    elif method == 'hyperimba':
        return 13
    else:
        raise NotImplementedError()


def get_score_offset(score):
    if score == 'acc':
        return 0
    elif score == 'bacc':
        return 1
    elif score == 'f1':
        return 2
    elif score == 'auc':
        return 3
    else:
        raise NotImplementedError()


def get_dataset_offset(dataset):
    if dataset == 'Cora_100':
        return (4, 2)
    elif dataset == 'Cora_20':
        return (4, 6)
    elif dataset == 'Cora_1':
        return (4, 10)
    elif dataset == 'CiteSeer_100':
        return (4, 14)
    elif dataset == 'CiteSeer_20':
        return (4, 18)
    elif dataset == 'CiteSeer_1':
        return (4, 22)
    elif dataset == 'PubMed_100':
        return (4, 26)
    elif dataset == 'PubMed_20':
        return (4, 30)
    elif dataset == 'PubMed_1':
        return (4, 34)
    elif dataset == 'chameleon_100':
        return (4, 38)
    elif dataset == 'chameleon_20':
        return (4, 42)
    elif dataset == 'chameleon_1':
        return (4, 46)
    elif dataset == 'squirrel_100':
        return (4, 50)
    elif dataset == 'squirrel_20':
        return (4, 54)
    elif dataset == 'squirrel_1':
        return (4, 58)
    elif dataset == 'Actor_100':
        return (4, 62)
    elif dataset == 'Actor_20':
        return (4, 66)
    elif dataset == 'Actor_1':
        return (4, 70)
    elif dataset == 'Wisconsin_100':
        return (4, 74)
    elif dataset == 'Wisconsin_20':
        return (4, 78)
    elif dataset == 'Wisconsin_1':
        return (4, 82)
    elif dataset == 'Amazon-Photo_100':
        return (22, 2)
    elif dataset == 'Amazon-Photo_20':
        return (22, 6)
    elif dataset == 'Amazon-Photo_1':
        return (22, 10)
    elif dataset == 'Amazon-Computers_100':
        return (22, 14)
    elif dataset == 'Amazon-Computers_20':
        return (22, 18)
    elif dataset == 'Amazon-Computers_1':
        return (22, 22)
    elif dataset == 'Coauthor-CS':
        return (22, 14)
    elif dataset == 'ogbn-arxiv':
        return (22, 26)
    else:
        raise NotImplementedError()
    
records_file_list = [
    # ('Photo', True),
    # ('Computers', True),
    # ('CS', True),
    ('vanilla', True),
    ('drgcn', True),
    ('smote', True),
    ('imgagn', True),
    ('ens', True),
    ('tam', True),
    ('sha', True),
    ('vanilla_smote_a', True),
    ('imgagn_ens_a', True),
    ('tam_sha_a', True),
    ('drgcn_a', True),
    # ('smote', True),
    # ('smote2', True),
    # ('ccp', False),
]

records = []

for records_file_item in records_file_list:
    print(records_file_item)
    if records_file_item[1]:
        records_file = 'records/' + records_file_item[0] + '_gpu.json'
        # records_file = 'records/records_gpu_' + records_file_item[0] + '.json'
    else:
        records_file = 'records/' + records_file_item[0] + '.json'
        # records_file = 'records/records_' + records_file_item[0] + '.json'
    if os.path.exists(records_file):
        with open(records_file) as f:
            records_this = json.load(f)
            records += records_this

benchmarks = dict()

for method in method_list:
    for dataset in dataset_list:
        benchmarks[(method, dataset)] = list()

for record in records:
    benchmarks[(record['method'], record['dataset'])].append((record['acc'], record['bacc'], record['f1'], record['auc']))

template_filename = 'Imbalance Benchmark Template.xlsx'
filename = 'Imbalance Benchmark.xlsx'

os.system('cp \'' + template_filename + '\' \'' + filename + '\'')
workbook = load_workbook(filename=filename)
ws4 = workbook["Node-level class imbalance "]

for method in method_list:
    for dataset in dataset_list:
        if len(benchmarks[(method, dataset)]) == len(seed_list):
            scores = dict()
            for score in score_list:
                scores[score] = list()
            for acc, bacc, f1, auc in benchmarks[(method, dataset)]:
                scores['acc'].append(acc)
                scores['bacc'].append(bacc)
                scores['f1'].append(f1)
                scores['auc'].append(auc)
            for score in score_list:
                mean_score = statistics.mean(scores[score])
                stdev_score = statistics.stdev(scores[score])
                row = 0
                column = 0
                row += get_method_offset(method=method)
                column += get_score_offset(score=score)
                row += get_dataset_offset(dataset=dataset)[0]
                column += get_dataset_offset(dataset=dataset)[1]
                ws4.cell(row=row, column=column).value = '%.3f (%.3f)' % (mean_score, stdev_score)
                # ws4.cell(row=row, column=column).value = mean_score.__str__() + ' (' + stdev_score.__str__() + ')'
            
workbook.save(filename=filename)
