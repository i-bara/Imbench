from benchmark import *
import json
from sklearn.metrics import silhouette_score
import torch
import shutil
import openpyxl
import datetime
import pickle


if __name__ == '__main__':
    name = args.name

    suffix = ''

    if args.gpu:
        suffix += '_gpu'

    if args.debug:
        raise RuntimeError('Cannot analyze with --debug option')

    records_file = os.path.join('records/', f'{name}{suffix}.json')
    with open(records_file) as f:
        records = json.load(f)
    
    # print(records)

    template_filename = 'Imbalance Benchmark Template.xlsx'
    filename = 'Imbalance Benchmark.xlsx'

    shutil.copyfile(template_filename, filename)
    workbook = openpyxl.load_workbook(filename=filename)
    ws4 = workbook["Node-level class imbalance "]

    l = ["vanilla", "drgcn", "dpgnn", "imgagn", "smote", "ens", "mixup", "lte4g", "tam", "topoauc", "sha"]
    l_dict = {y: x + 1 for x, y in enumerate(l)}
    d = [('Cora', 20.0), ('Actor', 20.0), ('ogbn-arxiv', 20.0)]
    d_dict = {y: x + 1 for x, y in enumerate(d)}

    analyze_dir = 'analyze/'
    if os.path.isdir(analyze_dir):
        shutil.rmtree(analyze_dir)
    os.mkdir(analyze_dir)

    for record in records:
        if record['dataset'] == 'Cora' and record['imb_ratio'] == 20.0:
            # shutil.copyfile(record['output'], os.path.join(analyze_dir, f'{record['method']}_{output}'))

            output = torch.load(record['output']).cpu().detach().numpy()
            y = torch.load(record['y']).cpu().detach().numpy()

            output_dict = dict()
            for i in range(output.shape[1]):
                output_i = output[y == i]
                output_dict[i] = output_i

            with open(os.path.join(analyze_dir, f'{record['method']}'), 'wb') as handle:
                pickle.dump(output_dict, handle)

            # print(f'{record['method']}: {output.shape}, {y.shape}')
            score = silhouette_score(output, y)
            ws4.cell(row=l_dict[record['method']], column=1).value = '%3f' % score

        t = datetime.datetime.strptime(record['time_erased'], "%H:%M:%S.%f")
        delta = datetime.timedelta(hours=t.hour, minutes=t.minute, seconds=t.second, microseconds=t.microsecond)
        ws4.cell(row=l_dict[record['method']] + 3, column=3 * d_dict[(record['dataset'], record['imb_ratio'])]).value = '%.3f' % delta.total_seconds()
        ws4.cell(row=l_dict[record['method']] + 3, column=3 * d_dict[(record['dataset'], record['imb_ratio'])] + 1).value = '%.3f' % (record['max_memory_allocated'] / 1024 / 1024)
        

    workbook.save(filename=filename)
