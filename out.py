import os
import json
import argparse
import openpyxl
import statistics
from functools import cmp_to_key


def parse_args():
    parser = argparse.ArgumentParser()
    
    parser.add_argument('name')
    parser.add_argument('--gpu', action='store_true',
                        help='use gpu')
    args = parser.parse_args()

    return args


args = parse_args()

lists = {
    'method': ['vanilla', 'drgcn', 'smote', 'imgagn', 'ens', 'tam', 'lte4g', 'topoauc', 'sann', 'sha', 'renode', 'pastel', 'hyperimba'],
    'dataset': ['Cora', 'CiteSeer', 'PubMed', 'chameleon', 'squirrel', 'Actor', 'Wisconsin', 'Amazon-Photo', 'Amazon-Computers', 'ogbn-arxiv'],
    'imb_ratio': [100.0, 20.0, 1.0],
    'seed': [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000],
}


def compare_other(key, x, y):
    if key in lists.keys():
        try:
            xx = lists[key].index(x)
            try:
                yy = lists[key].index(y)
                return xx > yy
            except:
                return False
        except ValueError:
            try:
                yy = lists[key].index(y)
                return True
            except:
                return x > y
    else:
        raise KeyError


compares = {
    'imb_ratio': lambda x, y: x < y,
    'seed': lambda x, y: x > y,
}

score_list = ['acc', 'bacc', 'f1', 'auc']

records_file_list = [
    (args.name, args.gpu),
]

records = []

for records_file_item in records_file_list:
    print(records_file_item)
    if records_file_item[1]:
        records_file = 'records/' + records_file_item[0] + '_gpu.json'
        # records_file = 'records/records_gpu_' + records_file_item[0] + '.json'
    else:
        records_file = 'records/' + records_file_item[0] + '.json'
        # records_file = 'records/records_' + records_file_item[0] + '.json'
    if os.path.exists(records_file):
        with open(records_file) as f:
            records_this = json.load(f)
            records += records_this

benchmarks = dict()

rows = ['method']
cols = ['dataset', 'imb_ratio', 'score']

keys = rows + cols
keys.remove('score')

for record in records:
    # benchmark = list()
    # for key in keys:
    #     benchmark.append(record[key])
    benchmark = tuple(map(lambda key: record[key], keys))
    if benchmark not in benchmarks.keys():
        benchmarks[benchmark] = list()
    benchmarks[benchmark].append(record)


def make_comparator(less_than):
    def compare(x, y):
        if less_than(x, y):
            return 1
        elif less_than(y, x):
            return -1
        else:
            return 0
    return compare


recorded_lists = dict()
for key in keys:
    key_set = set()
    compare = compares[key] if key in compares.keys() else lambda x, y: compare_other(key=key, x=x, y=y)
    for benchmark in benchmarks.keys():
        benchmark_keys = {keys[i]: benchmark[i] for i in range(len(keys))}
        key_set.add(benchmark_keys[key])
    recorded_lists[key] = sorted(list(key_set), key=cmp_to_key(make_comparator(compare)))
recorded_lists['score'] = score_list

# template_filename = 'Imbalance Benchmark Template.xlsx'
filename = f'{args.name}.xlsx'

# os.system('cp \'' + template_filename + '\' \'' + filename + '\'')
# workbook = load_workbook(filename=filename)
# ws4 = workbook["Node-level class imbalance "]

print(os.path.isfile(filename))

if os.path.isfile(filename):
    os.remove(filename)
print(os.path.isfile(filename))
workbook = openpyxl.Workbook()
sheetname = workbook.sheetnames[0]
sheet = workbook[sheetname]

for benchmark in benchmarks.keys():
    benchmark_keys = {keys[i]: benchmark[i] for i in range(len(keys))}
    records = benchmarks[benchmark]
    scores = dict()
    for score in score_list:
        scores[score] = list()
    for record in records:
        for score in score_list:
            scores[score].append(record[score])
    for score in score_list:
        mean_score = statistics.mean(scores[score])
        stdev_score = statistics.stdev(scores[score])
        n_row = 0
        n_col = 0
        for row in rows:
            n_row *= len(recorded_lists[row])
            n_row += recorded_lists[row].index(score if row == 'score' else benchmark_keys[row])
        for col in cols:
            n_col *= len(recorded_lists[col])
            n_col += recorded_lists[col].index(score if col == 'score' else benchmark_keys[col])
        sheet.cell(row=n_row+len(cols)+1, column=n_col+len(rows)+1).value = '%.3f (%.3f)' % (mean_score, stdev_score)

row_left = 1
row_right = 1
for row in rows:
    row_right *= len(recorded_lists[row])
for row_idx, row in enumerate(rows):
    row_center = len(recorded_lists[row])
    row_right //= row_center
    for i in range(row_left):
        for j in range(row_center):
            sheet.cell(row=((i * row_center) + j)*row_right+len(cols)+1, column=row_idx+1).value = recorded_lists[row][j]
            sheet.merge_cells(start_row=((i * row_center) + j)*row_right+len(cols)+1, start_column=row_idx+1, end_row=((i * row_center) + j)*row_right+len(cols)+row_right, end_column=row_idx+1)
    row_left *= row_center

col_left = 1
col_right = 1
for col in cols:
    col_right *= len(recorded_lists[col])
for col_idx, col in enumerate(cols):
    col_center = len(recorded_lists[col])
    col_right //= col_center
    for i in range(col_left):
        for j in range(col_center):
            sheet.merge_cells(start_row=col_idx+1, start_column=((i * col_center) + j)*col_right+len(rows)+1, end_row=col_idx+1, end_column=((i * col_center) + j)*col_right+len(rows)+col_right)
            sheet.cell(row=col_idx+1, column=((i * col_center) + j)*col_right+len(rows)+1).value = recorded_lists[col][j]
            
    col_left *= col_center

workbook.save(filename=filename)
