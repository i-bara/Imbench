import os
import re
import json
import datetime
import pandas as pd
from openpyxl import load_workbook
import argparse


def parse_args():
    parser = argparse.ArgumentParser()
    # Method
    parser.add_argument('--gpu', action='store_true',
                        help='use gpu')
    args = parser.parse_args()

    return args


args = parse_args()

method_list = ['vanilla', 'drgcn', 'smote', 'imgagn', 'ens', 'tam', 'lte4g', 'sann', 'sha', 'renode', 'pastel', 'hyperimba']
score_list = ['acc', 'bacc', 'macrof1']
dataset_list = ['Cora_100', 'Cora_20', 'CiteSeer_100', 'CiteSeer_20', 'PubMed_100', 'PubMed_20', 'Amazon-Photo', 'Amazon-Computers', 'Coauthor-CS']
seed_list = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]

def get_method_offset(method):
    if method == 'vanilla':
        return 0
    elif method == 'drgcn':
        return 1
    elif method == 'smote':
        return 2
    elif method == 'imgagn':
        return 3
    elif method == 'ens':
        return 4
    elif method == 'tam':
        return 5
    elif method == 'lte4g':
        return 6
    elif method == 'sann':
        return 7
    elif method == 'sha':
        return 8
    elif method == 'renode':
        return 10
    elif method == 'pastel':
        return 11
    elif method == 'hyperimba':
        return 12
    else:
        raise NotImplementedError()


def get_score_offset(score):
    if score == 'acc':
        return 0
    elif score == 'bacc':
        return 1
    elif score == 'macrof1':
        return 2
    else:
        raise NotImplementedError()


def get_dataset_offset(dataset):
    if dataset == 'Cora_100':
        return (4, 2)
    elif dataset == 'Cora_20':
        return (4, 5)
    elif dataset == 'CiteSeer_100':
        return (4, 9)
    elif dataset == 'CiteSeer_20':
        return (4, 11)
    elif dataset == 'PubMed_100':
        return (4, 14)
    elif dataset == 'PubMed_20':
        return (4, 17)
    else:
        raise NotImplementedError()


def config(method, dataset, seed):
    method_config = '--method ' + method
    if dataset == 'Cora_100':
        dataset_config = '--dataset Cora --imb_ratio 100'
    elif dataset == 'Cora_20':
        dataset_config = '--dataset Cora --imb_ratio 20'
    elif dataset == 'CiteSeer_100':
        dataset_config = '--dataset CiteSeer --imb_ratio 100'
    elif dataset == 'CiteSeer_20':
        dataset_config = '--dataset CiteSeer --imb_ratio 20'
    elif dataset == 'PubMed_100':
        dataset_config = '--dataset PubMed --imb_ratio 100'
    elif dataset == 'PubMed_20':
        dataset_config = '--dataset PubMed --imb_ratio 20'
    elif dataset in ['Amazon-Photo', 'Amazon-Computers', 'Coauthor-CS']:
        dataset_config = '--dataset ' + dataset
    else:
        raise NotImplementedError()
    return method_config + ' ' + dataset_config + ' --seed ' + seed.__str__()


# os.system("ping baidu.com > ping.txt")


def experiment(method, dataset, seed):
    done = False
    for record in records:
        if record['method'] == method and record['dataset'] == dataset and record['seed'] == seed:
            done = True

    if not done:
        begin_datetime = datetime.datetime.now()
        if args.gpu:
            os.system("python main.py " + config(method=method, dataset=dataset, seed=seed) + " > cache4.txt")
        else:
            os.system("python main.py " + config(method=method, dataset=dataset, seed=seed) + " --device cpu > cache4.txt")
        end_datetime = datetime.datetime.now()

        with open("cache4.txt") as f:
            info = f.readline()
            match = re.match('acc: ([\\d\\.]*), bacc: ([\\d\\.]*), f1: ([\\d\\.]*)', info)
            if match is not None:
                acc, bacc, f1 = tuple(map(lambda x: float(x), match.groups()))
                record = dict()
                record['begin_datetime'] = begin_datetime.__str__()
                record['end_datetime'] = end_datetime.__str__()
                record['time_erased'] = (end_datetime - begin_datetime).__str__()
                record['method'] = method
                record['dataset'] = dataset
                record['seed'] = seed
                record['acc'] = acc
                record['bacc'] = bacc
                record['f1'] = f1
                print(record)
                records.append(record)

        os.system("rm cache4.txt")
    
    with open(records_file, 'w+') as f:
        json.dump(records, f, indent=4)


# Begin

if args.gpu:
    records_file = 'records_gpu4.json'
else:
    records_file = 'records4.json'

if os.path.exists(records_file):
    with open(records_file) as f:
        records = json.load(f)
else:
    records = []

methods = ['vanilla', 'smote', 'ens', 'sha']
# methods = ['vanilla', 'ens', 'sha']
datasets = ['Amazon-Computers']
seeds = seed_list

print(f'''
methods = {methods}
datasets = {datasets}
seeds = {seeds}
''')

# methods = ['vanilla', 'ens', 'sha']
# datasets = ['Cora_100']
# seeds = [100, 200]

for method in methods:
    for dataset in datasets:
        for seed in seeds:
            experiment(method=method, dataset=dataset, seed=seed)

exit()

filename = 'Imbalance Benchmark.xlsx'

#load excel file
workbook = load_workbook(filename=filename)

#Pick the sheet "new_sheet"
ws4 = workbook["Sheet1"]

#modify the desired cell
ws4.cell(row = 1, column = 3).value = 'Old Price'

print(ws4.cell(row = 1, column = 4).value)

#save the file
workbook.save(filename=filename)
