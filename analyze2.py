from benchmark import *
import json
from sklearn.metrics import silhouette_score
import torch
import shutil
import openpyxl
import datetime
import pickle
from sklearn.metrics import f1_score
import numpy as np


if __name__ == '__main__':
    name = args.name

    suffix = ''

    if args.gpu:
        suffix += '_gpu'

    if args.debug:
        raise RuntimeError('Cannot analyze with --debug option')

    records_file = os.path.join('records/', f'{name}{suffix}.json')
    with open(records_file) as f:
        records = json.load(f)
    
    # print(records)

    template_filename = 'Imbalance Benchmark Template.xlsx'
    filename = 'Imbalance Benchmark.xlsx'

    shutil.copyfile(template_filename, filename)
    workbook = openpyxl.load_workbook(filename=filename)
    workbook = openpyxl.Workbook()
    # workbook.create_sheet("Node-level class imbalance ")
    # ws4 = workbook["Node-level class imbalance "]
    ws4 = workbook["Sheet"]

    l = ["vanilla", "drgcn", "dpgnn", "imgagn", "smote", "ens", "mixup", "lte4g", "tam", "topoauc", "sha", "renode"]
    l_dict = {y: x + 1 for x, y in enumerate(l)}
    d = [('Cora', 20.0), ('Actor', 20.0), ('ogbn-arxiv', 20.0)]
    # d = [("CiteSeer", 20.0), ("PubMed", 20.0), ("chameleon", 20.0), ("squirrel", 20.0), ("Amazon-Photo", 20.0), ("Amazon-Computers", 20.0)]
    d_dict = {y: x + 1 for x, y in enumerate(d)}

    analyze_dir = 'analyze/'
    if os.path.isdir(analyze_dir):
        shutil.rmtree(analyze_dir)
    os.mkdir(analyze_dir)

    for record in records:
        if record['imb_ratio'] == 20.0:
            # shutil.copyfile(record['output'], os.path.join(analyze_dir, f'{record['method']}_{output}'))

            # .max(1)[1] to f1
            output = torch.load(record['output'])[torch.load(record['test'])].cpu().detach().numpy()
            y_pred = torch.load(record['output'])[torch.load(record['test'])].max(1)[1].cpu().detach().numpy()
            y = torch.load(record['y'])[torch.load(record['test'])].cpu().detach().numpy()
            
            y_train = torch.load(record['y'])[torch.load(record['train'])].cpu().detach().numpy()

            output_dict = dict()
            for i in range(output.shape[1]):
                output_i = output[y == i]
                output_dict[i] = output_i

            with open(os.path.join(analyze_dir, f'{record['method']}'), 'wb') as handle:
                pickle.dump(output_dict, handle)

            # print(f'{record['method']}: {output.shape}, {y.shape}')
            
            f1 = f1_score(y, y_pred, average=None)
            print(record['method'], f1)
            
            cls_num = [y_train[y_train == c].shape[0] for c in range(f1.shape[0])]
            cls_num_sum = sum(cls_num)
            for c in range(f1.shape[0]):
                ws4.cell(row=4, column=d_dict[(record['dataset'], record['imb_ratio'])] * 30 - 28 + c).value = f"{cls_num[c]} ({'%.1f%%' % (cls_num[c] / cls_num_sum * 100)})"
            # ws4.cell(row=4, column=d_dict[(record['dataset'], record['imb_ratio'])] * 30 - 28 + f1.shape[0]).value = 
            ws4.cell(row=4, column=d_dict[(record['dataset'], record['imb_ratio'])] * 30 - 29).value = f"{record['dataset']}(lt={record['imb_ratio']})"
            ws4.cell(row=l_dict[record['method']] + 4, column=d_dict[(record['dataset'], record['imb_ratio'])] * 30 - 29).value = f"{record['method']}"
            for c, score in enumerate(f1):
                ws4.cell(row=l_dict[record['method']] + 4, column=d_dict[(record['dataset'], record['imb_ratio'])] * 30 - 28 + c).value = '%.3f' % (score * 100)
                
            ws4.cell(row=l_dict[record['method']] + 4, column=d_dict[(record['dataset'], record['imb_ratio'])] * 30  - 28+ f1.shape[0]).value = '%.3f' % (np.mean(f1) * 100)
            
            
            # score = silhouette_score(output, y)
            # ws4.cell(row=l_dict[record['method']], column=d_dict[record['dataset'] * 8 + 1]).value = '%3f' % score
            # ws4.cell(row=l_dict[record['method']], column=2).value = '%3f' % record['f1']
            # print(output)
            # print(record['method'], score)

        # t = datetime.datetime.strptime(record['time_erased'], "%H:%M:%S.%f")
        # delta = datetime.timedelta(hours=t.hour, minutes=t.minute, seconds=t.second, microseconds=t.microsecond)
        # ws4.cell(row=l_dict[record['method']] + 3, column=3 * d_dict[(record['dataset'], record['imb_ratio'])]).value = '%.3f' % delta.total_seconds()
        # ws4.cell(row=l_dict[record['method']] + 3, column=3 * d_dict[(record['dataset'], record['imb_ratio'])] + 1).value = '%.3f' % (record['max_memory_allocated'] / 1024 / 1024)
        

    workbook.save(filename=filename)
